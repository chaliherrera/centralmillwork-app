import openpyxl, re, datetime, sys

# Ruta del Excel: 1er argumento de línea de comandos, o el default de siempre.
SRC = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\chali\OneDrive - Central Millwork\Desktop\UTILITIES\Master.Sched.xlsx"
OUT = "import_ing.sql"

# Catálogo canónico: (clave, nombre, hito, tipico, min, max, orden, [aliases])
TIPOS = [
 ("field_measurements","Field Measurements","E-03",1,1,1,10,["field measurements","field measurement","vif","field measurements "]),
 ("samples","Samples Process","E-05",12,10,15,20,["samples process","samples","architect/designer review and samp"]),
 ("shop_drawings","Shop Drawings Process","E-06",10,5,15,30,["shop drawings process","shop drawings","shop drawings process rev0","shop drawings process","sd update/final production set","update/final production set","sd update / final production set","sd update/final prod","sd update/final production set "]),
 ("client_review","Architect/Designer Review","E-07",10,1,30,40,["architect/designer review drawings","architect/designer review","architec/designer review drawings","meeting with designer to review pr","meeting with vantage to review pro","architect/designer review and samp"]),
 ("release","Release to Production","E-10",0,0,0,50,["release to production","release to productio","release to production "]),
 ("cnc","CNC Engineering","E-11",5,1,8,60,["cnc engineering"]),
 ("long_leads","Long Lead Material Procurement","M-03",20,15,25,25,["long lead time material procuremen","long lead time material procurement"]),
 ("material_proc","Material Procurement","M-04",5,1,10,26,["material procurement"]),
 ("fabrication","Millwork Fabrication","P-05",15,5,25,70,["millwork fabrication"]),
 ("installation","Millwork Installation","I-04",7,1,15,80,["millwork installation","milwork installation"]),
 ("shipment","Millwork Shipment","S-04",0,0,0,75,["milwork shipment","millwork shipment"]),
 ("stone_measure","Stone Countertop Measuring","E-03",1,1,1,81,["stone countertop measuring and tem","stone countertop measuring and temp"]),
 ("stone_fab","Stone Countertops Fabrication","P-05",7,5,10,82,["stone countertops fabrication","stone countertop fabrication"]),
 ("stone_install","Stone Countertops Installation","I-04",2,1,3,83,["stone countertops installation"]),
]
alias2clave={}
for c,n,h,t,mn,mx,o,al in TIPOS:
    alias2clave[n.lower().strip()]=c
    for a in al: alias2clave[a.lower().strip()]=c

def resolve(name):
    return alias2clave.get((name or "").lower().strip())

def sq(v):
    if v is None: return "NULL"
    return "'" + str(v).replace("'","''") + "'"
def num(v):
    return "NULL" if v is None else str(v)
def dat(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return "'"+v.strftime("%Y-%m-%d")+"'"
    return "NULL"
def pdur(v):
    if v is None: return 1
    if isinstance(v,(int,float)): return v
    m=re.match(r'([\d.]+)',str(v)); return float(m.group(1)) if m else 1

# Parsea la notación de predecesores del Smartsheet: "93FS+6d", "99", "12,15", "8SS-2d".
# El número es la FILA de Smartsheet (1-indexada, arranca en la fila-proyecto); el xlsx
# tiene header en la fila 1, así que fila-Smartsheet N == fila-xlsx (N+1). Devuelve una
# lista de (xlsx_row_destino, tipo, lag_dias). Verificado con casos reales del Excel.
DEP_RE=re.compile(r'^\s*(\d+)\s*(FS|SS|FF|SF)?\s*([+-]\s*\d+)?\s*d?\s*$', re.I)
def parse_preds(v):
    if v in (None,''): return []
    out=[]
    for tok in str(v).replace(';',',').split(','):
        m=DEP_RE.match(tok)
        if not m: continue
        ss_row=int(m.group(1))
        tipo=(m.group(2) or 'FS').upper()
        lag=int(m.group(3).replace(' ','')) if m.group(3) else 0
        out.append((ss_row+1, tipo, lag))   # +1: Smartsheet -> xlsx
    return out

wb=openpyxl.load_workbook(SRC,data_only=True)
ws=wb["Master.Sched"]
H=[c.value for c in ws[1]]; idx={h:i for i,h in enumerate(H)}
def g(r,name):
    i=idx.get(name); return r[i] if i is not None and i<len(r) else None
proj_re=re.compile(r'^\s*(\d{2}-\d{3})')
prefix_re=re.compile(r'^\s*\d{2}-\d{3}\s+')   # para quitar el codigo de tareas prefijadas

lines=[]
lines.append("BEGIN;")
# IMPORTANTE: NO borrar todo. Solo refrescamos lo que vino del Excel (origen='import_excel').
# Las reservas (origen='reserva') y las tareas manuales se PRESERVAN. Las deps de las filas
# borradas caen por FK ON DELETE CASCADE. (Antes: TRUNCATE — borraba las reservas. Bug corregido.)
lines.append("DELETE FROM ing_tareas WHERE origen='import_excel';")
# Los encabezados de proyecto (fecha fija) también se refrescan. Los manuales se preservan.
lines.append("DELETE FROM ing_proyectos WHERE origen='import_excel';")
# seed catálogo (idempotente)
for c,n,h,t,mn,mx,o,al in TIPOS:
    arr="ARRAY[" + ",".join(sq(a) for a in al) + "]::text[]"
    lines.append(f"INSERT INTO ing_tarea_tipos (clave,nombre,hito_codigo,dur_dias_tipico,dur_dias_min,dur_dias_max,orden,aliases) "
                 f"VALUES ({sq(c)},{sq(n)},{sq(h)},{t},{mn},{mx},{o},{arr}) "
                 f"ON CONFLICT (clave) DO UPDATE SET nombre=EXCLUDED.nombre,hito_codigo=EXCLUDED.hito_codigo,"
                 f"dur_dias_tipico=EXCLUDED.dur_dias_tipico,dur_dias_min=EXCLUDED.dur_dias_min,dur_dias_max=EXCLUDED.dur_dias_max,aliases=EXCLUDED.aliases;")
# Regla día-por-ítem (shop drawings ≈ 1 día/ítem). Se setea en cada import para que
# no se pierda si el catálogo se re-inserta (la migración 057 no alcanza si corre antes
# de que existan las filas del catálogo).
lines.append("UPDATE ing_tarea_tipos SET dias_por_item = 1.0 WHERE clave = 'shop_drawings';")

cur_proj=None; cur_code=None; cur_phase=None; ntask=0; nproj=0; deps=[]
for rn,r in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
    name=(g(r,'Project Name') or '').strip()
    if not name: continue
    # El formato nuevo del Excel prefija TODAS las filas con el codigo ("25-562 Field
    # measurements"). El encabezado de proyecto = primera fila con un codigo NUEVO; las
    # tareas prefijadas repiten el codigo del proyecto actual, asi que NO son proyectos.
    m=proj_re.match(name); code=m.group(1) if m else None
    if code and code != cur_code:
        cur_code=code; cur_proj=name; cur_phase=None; nproj+=1
        # Encabezado del proyecto: la FECHA FIJA (Finish) + inicio + total + estado.
        lines.append(
          "INSERT INTO ing_proyectos (proyecto_ext,fecha_inicio,fecha_entrega,dur_total_dias,status_ext,origen) VALUES ("
          f"{sq(cur_proj)},{dat(g(r,'Start'))},{dat(g(r,'Finish'))},{pdur(g(r,'Duration'))},"
          f"{sq((g(r,'Status') or '').strip() or None)},'import_excel') "
          "ON CONFLICT (proyecto_ext) DO UPDATE SET fecha_inicio=EXCLUDED.fecha_inicio,fecha_entrega=EXCLUDED.fecha_entrega,"
          "dur_total_dias=EXCLUDED.dur_total_dias,status_ext=EXCLUDED.status_ext,updated_at=NOW();")
        continue
    # Tarea (o fase): quita el prefijo de codigo si lo trae.
    name=prefix_re.sub('',name).strip()
    if name.lower().startswith('phase'):
        cur_phase=name; continue
    assigned=(g(r,'Assigned To') or '').strip() or None
    dur=pdur(g(r,'Duration'))
    alloc=g(r,'Allocation %'); alloc=alloc if isinstance(alloc,(int,float)) else None
    status=(g(r,'Status') or '').strip() or None
    preds=g(r,'Predecessors'); preds=str(preds) if preds not in (None,'') else None
    com=(g(r,'Coments') or g(r,'Comments') or '').strip() or None
    clave=resolve(name)
    tipo_sql = f"(SELECT id FROM ing_tarea_tipos WHERE clave={sq(clave)})" if clave else "NULL"
    ext=f"xlsx-row-{rn}"
    lines.append(
      "INSERT INTO ing_tareas (proyecto_ext,fase,tipo_id,nombre,asignado_nombre,allocation_pct,dur_dias,"
      "fecha_inicio,fecha_fin,status_ext,predecesores_ext,comentario,origen,external_ref) VALUES ("
      f"{sq(cur_proj)},{sq(cur_phase)},{tipo_sql},{sq(name)},{sq(assigned)},{num(alloc if alloc is not None else 1)},"
      f"{dur},{dat(g(r,'Start'))},{dat(g(r,'Finish'))},{sq(status)},{sq(preds)},{sq(com)},'import_excel',{sq(ext)}) "
      "ON CONFLICT (external_ref) DO UPDATE SET proyecto_ext=EXCLUDED.proyecto_ext,fase=EXCLUDED.fase,"
      "tipo_id=EXCLUDED.tipo_id,nombre=EXCLUDED.nombre,asignado_nombre=EXCLUDED.asignado_nombre,"
      "allocation_pct=EXCLUDED.allocation_pct,dur_dias=EXCLUDED.dur_dias,fecha_inicio=EXCLUDED.fecha_inicio,"
      "fecha_fin=EXCLUDED.fecha_fin,status_ext=EXCLUDED.status_ext,comentario=EXCLUDED.comentario,updated_at=NOW();")
    ntask+=1
    for tgt_row, tipo, lag in parse_preds(preds):
        deps.append((rn, tgt_row, tipo, lag))

# Dependencias (aristas editables): resuelve external_ref -> id. Las que apunten a
# filas que no son tareas (proyecto/fase) no matchean y se saltan solas. Las deps de
# tareas import_excel ya se borraron por CASCADE al borrar sus tareas arriba.
for src_row, tgt_row, tipo, lag in deps:
    lines.append(
      "INSERT INTO ing_tarea_deps (tarea_id,depende_de_id,tipo,lag_dias) "
      f"SELECT t.id, d.id, {sq(tipo)}, {lag} FROM ing_tareas t JOIN ing_tareas d ON true "
      f"WHERE t.external_ref='xlsx-row-{src_row}' AND d.external_ref='xlsx-row-{tgt_row}' "
      "ON CONFLICT (tarea_id,depende_de_id) DO UPDATE SET tipo=EXCLUDED.tipo, lag_dias=EXCLUDED.lag_dias;")
# Anti doble-conteo: si el Excel trae las tareas reales de un proyecto que tenia una
# reserva provisional (sin confirmar), esas reservas ya no aplican -> estado='na'.
# Las confirmadas por el PM se preservan (son compromiso real).
lines.append("UPDATE ing_tareas SET estado='na', updated_at=NOW() "
             "WHERE origen='reserva' AND reserva_confirmada_at IS NULL "
             "AND proyecto_ext IN (SELECT DISTINCT proyecto_ext FROM ing_tareas WHERE origen='import_excel');")
lines.append("COMMIT;")
open(OUT,"w",encoding="utf-8").write("\n".join(lines))
print(f"SQL generado: {OUT} | proyectos={nproj} tareas={ntask} dependencias={len(deps)}")
